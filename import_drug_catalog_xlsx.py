from pathlib import Path
import sqlite3
import re
import sys
import openpyxl

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "safepill.db"
DEFAULT_XLSX = BASE_DIR / "의약품 허가 품목 정보.xlsx"


def normalize_text(s: str) -> str:
    if not s:
        return ""
    s = str(s)
    s = re.sub(r"\(.*?\)|\[.*?\]|\{.*?\}", " ", s)
    s = re.sub(r"(mg|g|ml|mcg|㎎|㎖|정|캡슐|환|포|병)\b", " ", s, flags=re.I)
    s = re.sub(r"[^0-9A-Za-z가-힣]+", "", s)
    return s.strip()


def clean_space(s: str) -> str:
    return " ".join(str(s or "").replace("_x000D_", "\n").split()).strip()


def parse_ingredients(main_text: str, raw_text: str):
    found = []

    def add(x: str):
        x = clean_space(x)
        x = re.sub(r"^[^가-힣A-Za-z]+", "", x)
        x = re.sub(r"[^0-9A-Za-z가-힣·ㆍ\-\+\(\)]+$", "", x)
        if not x:
            return
        if re.fullmatch(r"(EP|USP|KP|NF|JP|BP|생규|별첨규격.*|전과동|밀리그램|mg|g|mcg|mL|ml)", x, flags=re.I):
            return
        if not re.search(r"[가-힣A-Za-z]", x):
            return
        if x not in found:
            found.append(x)

    for part in re.split(r"[|;,\n]+", str(main_text or "")):
        m = re.search(r"\]([^\]|;,\n]+)", part)
        add(m.group(1) if m else part)

    for seg in str(raw_text or "").split(";"):
        pieces = [p.strip() for p in seg.split("|") if str(p).strip()]
        if len(pieces) >= 2:
            add(pieces[1])

    cleaned = []
    for x in found:
        x = re.sub(r"\(으\)로서.*$", "", x).strip()
        x = re.sub(r"\d+(?:\.\d+)?\s*(?:밀리그램|mg|g|mcg|mL|ml).*$", "", x, flags=re.I).strip()
        x = re.sub(r"\s+", " ", x).strip()
        if x and x not in cleaned:
            cleaned.append(x)
    return cleaned


def alias_variants(name: str):
    name = clean_space(name)
    variants = []

    def add(x: str):
        x = clean_space(x)
        if x and x not in variants:
            variants.append(x)

    add(name)
    no_paren = re.sub(r"\([^)]*\)", "", name).strip()
    add(no_paren)
    no_space = re.sub(r"\s+", "", no_paren)
    add(no_space)
    no_dose_keep_form = re.sub(r"((\d+(?:\.\d+)?)\s*(mg|g|mcg|ml|mL|㎎|㎖|밀리그램|그램))", "", no_paren, flags=re.I).strip()
    add(no_dose_keep_form)
    no_dose = re.sub(r"((\d+(?:\.\d+)?)\s*(mg|g|mcg|ml|mL|㎎|㎖|밀리그램|그램))", "", no_paren, flags=re.I)
    no_dose = re.sub(r"(정|캡슐|연질캡슐|시럽|현탁액|과립|액|주)$", "", no_dose).strip()
    add(no_dose)
    return [v for v in variants if len(v) >= 2]


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f"엑셀 파일을 찾을 수 없습니다: {xlsx_path}")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.executescript(
        """
        DROP TABLE IF EXISTS drug_catalog;
        DROP TABLE IF EXISTS drug_catalog_aliases;
        DROP TABLE IF EXISTS drug_catalog_ingredients;
        CREATE TABLE drug_catalog (
            product_id INTEGER PRIMARY KEY,
            product_name TEXT NOT NULL,
            product_name_norm TEXT NOT NULL,
            item_seq TEXT,
            company_name TEXT,
            otc_class TEXT,
            efficacy_text TEXT,
            raw_ingredient_text TEXT,
            main_ingredient_text TEXT
        );
        CREATE INDEX idx_drug_catalog_norm ON drug_catalog(product_name_norm);
        CREATE TABLE drug_catalog_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            alias_name TEXT NOT NULL,
            alias_norm TEXT NOT NULL
        );
        CREATE INDEX idx_drug_catalog_alias_norm ON drug_catalog_aliases(alias_norm);
        CREATE TABLE drug_catalog_ingredients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            ingredient_name TEXT NOT NULL,
            ingredient_norm TEXT NOT NULL
        );
        CREATE INDEX idx_drug_catalog_ing_product ON drug_catalog_ingredients(product_id);
        CREATE INDEX idx_drug_catalog_ing_norm ON drug_catalog_ingredients(ingredient_norm);
        """
    )
    conn.commit()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hmap = {h: i for i, h in enumerate(headers)}
    needed = ["품목명", "품목일련번호", "업체명", "전문일반", "효능효과", "원료성분", "주성분명"]

    product_rows = []
    alias_rows = []
    ing_rows = []
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = {k: row[hmap[k]] if hmap.get(k) is not None and hmap[k] < len(row) else None for k in needed}
        product_name = clean_space(vals["품목명"])
        if not product_name:
            continue
        product_norm = normalize_text(product_name)
        if not product_norm:
            continue
        count += 1
        pid = count
        item_seq = str(vals["품목일련번호"] or "").strip()
        company = clean_space(vals["업체명"])
        otc = clean_space(vals["전문일반"])
        efficacy = clean_space(vals["효능효과"])[:5000]
        raw_ing = clean_space(vals["원료성분"])
        main_ing = clean_space(vals["주성분명"])

        product_rows.append((pid, product_name, product_norm, item_seq, company, otc, efficacy, raw_ing, main_ing))
        for alias in alias_variants(product_name):
            alias_norm = normalize_text(alias)
            if alias_norm:
                alias_rows.append((pid, alias, alias_norm))
        for ing in parse_ingredients(main_ing, raw_ing):
            ing_norm = normalize_text(ing)
            if ing_norm:
                ing_rows.append((pid, ing, ing_norm))

        if count % 2000 == 0:
            cur.executemany("INSERT INTO drug_catalog(product_id, product_name, product_name_norm, item_seq, company_name, otc_class, efficacy_text, raw_ingredient_text, main_ingredient_text) VALUES (?,?,?,?,?,?,?,?,?)", product_rows)
            cur.executemany("INSERT INTO drug_catalog_aliases(product_id, alias_name, alias_norm) VALUES (?,?,?)", alias_rows)
            cur.executemany("INSERT INTO drug_catalog_ingredients(product_id, ingredient_name, ingredient_norm) VALUES (?,?,?)", ing_rows)
            conn.commit()
            product_rows, alias_rows, ing_rows = [], [], []
            print(f"imported {count}")

    if product_rows:
        cur.executemany("INSERT INTO drug_catalog(product_id, product_name, product_name_norm, item_seq, company_name, otc_class, efficacy_text, raw_ingredient_text, main_ingredient_text) VALUES (?,?,?,?,?,?,?,?,?)", product_rows)
        cur.executemany("INSERT INTO drug_catalog_aliases(product_id, alias_name, alias_norm) VALUES (?,?,?)", alias_rows)
        cur.executemany("INSERT INTO drug_catalog_ingredients(product_id, ingredient_name, ingredient_norm) VALUES (?,?,?)", ing_rows)
        conn.commit()

    cur.executescript(
        """
        DELETE FROM drug_catalog_aliases
        WHERE id NOT IN (
          SELECT MIN(id) FROM drug_catalog_aliases GROUP BY product_id, alias_norm
        );
        DELETE FROM drug_catalog_ingredients
        WHERE id NOT IN (
          SELECT MIN(id) FROM drug_catalog_ingredients GROUP BY product_id, ingredient_norm
        );
        """
    )
    conn.commit()
    conn.close()
    print(f"완료: {count}개 품목 적재")


if __name__ == "__main__":
    main()
